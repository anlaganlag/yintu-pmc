import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
import os

def add_supplier_shortage_summary():
    """
    读取现有PMC排产分析文件，添加按供应商汇总的缺料金额分析表
    """
    
    # 读取现有文件
    file_path = "PMC排产Sep01-Sep07订单ROI分析(1).xlsx"
    
    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return
    
    try:
        # 读取主要数据表
        df_main = pd.read_excel(file_path, sheet_name=0)  # 读取第一个sheet
        print(f"读取主表数据: {len(df_main)} 行")
        print("主表列名:", df_main.columns.tolist())
        
        # 检查是否有供应商和缺料金额相关列
        supplier_cols = [col for col in df_main.columns if '供应商' in col]
        shortage_cols = [col for col in df_main.columns if '缺料' in col and ('金额' in col or 'RMB' in col)]
        
        print(f"供应商相关列: {supplier_cols}")
        print(f"缺料金额相关列: {shortage_cols}")
        
        if not supplier_cols or not shortage_cols:
            print("未找到供应商或缺料金额列，显示前几行数据:")
            print(df_main.head())
            return
        
        # 使用第一个找到的供应商列和缺料金额列
        supplier_col = supplier_cols[0]
        shortage_col = shortage_cols[0]
        
        print(f"使用供应商列: {supplier_col}")
        print(f"使用缺料金额列: {shortage_col}")
        
        # 处理数据类型 - 确保缺料金额是数值型
        df_clean = df_main.copy()
        df_clean[shortage_col] = pd.to_numeric(df_clean[shortage_col], errors='coerce').fillna(0)
        
        # 过滤掉供应商为空的行
        df_clean = df_clean[df_clean[supplier_col].notna() & (df_clean[supplier_col] != '')]
        
        print(f"清理后数据行数: {len(df_clean)}")
        print(f"总缺料金额: ¥{df_clean[shortage_col].sum():,.2f}")
        
        # 按供应商汇总缺料金额
        supplier_summary = df_clean.groupby(supplier_col).agg({
            shortage_col: ['sum', 'count'],
            '生产订单' if '生产订单' in df_clean.columns else df_clean.columns[0]: 'nunique'
        }).reset_index()
        
        # 重命名列
        supplier_summary.columns = [
            '供应商名称',
            '缺料金额合计(RMB)', 
            '缺料物料数量',
            '涉及订单数量'
        ]
        
        # 按缺料金额降序排列
        supplier_summary = supplier_summary.sort_values('缺料金额合计(RMB)', ascending=False)
        
        # 计算占比
        total_shortage = supplier_summary['缺料金额合计(RMB)'].sum()
        supplier_summary['占总缺料比例(%)'] = (supplier_summary['缺料金额合计(RMB)'] / total_shortage * 100).round(2)
        
        # 添加累计占比
        supplier_summary['累计占比(%)'] = supplier_summary['占总缺料比例(%)'].cumsum().round(2)
        
        # 格式化金额
        supplier_summary['缺料金额合计(RMB)'] = supplier_summary['缺料金额合计(RMB)'].round(2)
        
        print(f"\n供应商缺料汇总 (总计: ¥{total_shortage:,.2f}):")
        print(supplier_summary.to_string(index=False))
        
        # 加载现有工作簿
        book = load_workbook(file_path)
        
        # 创建新的工作表
        if '供应商缺料汇总' in book.sheetnames:
            del book['供应商缺料汇总']
        
        ws_summary = book.create_sheet('供应商缺料汇总')
        
        # 添加标题
        ws_summary.append(['9月第一周排产订单 - 供应商缺料金额汇总分析'])
        ws_summary.append([f'统计时间: 2025-09-01', f'总缺料金额: ¥{total_shortage:,.2f}'])
        ws_summary.append([])  # 空行
        
        # 添加数据
        for r in dataframe_to_rows(supplier_summary, index=False, header=True):
            ws_summary.append(r)
        
        # 设置样式
        # 标题样式
        title_cell = ws_summary['A1']
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        ws_summary.merge_cells('A1:G1')
        
        # 统计信息样式
        for cell in ws_summary[2]:
            if cell.value:
                cell.font = Font(size=10, italic=True)
        
        # 表头样式
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        for cell in ws_summary[4]:
            if cell.value:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
        
        # 数据格式化
        for row in ws_summary.iter_rows(min_row=5, max_row=ws_summary.max_row):
            # 金额列 (B列) 格式化为货币
            if row[1].value and isinstance(row[1].value, (int, float)):
                row[1].number_format = '#,##0.00'
            
            # 百分比列格式化
            if len(row) > 4 and row[4].value and isinstance(row[4].value, (int, float)):
                row[4].number_format = '0.00%'
                row[4].value = row[4].value / 100
            
            if len(row) > 5 and row[5].value and isinstance(row[5].value, (int, float)):
                row[5].number_format = '0.00%'  
                row[5].value = row[5].value / 100
        
        # 调整列宽
        from openpyxl.utils import get_column_letter
        column_widths = [20, 18, 12, 12, 15, 15]
        for i, width in enumerate(column_widths, 1):
            col_letter = get_column_letter(i)
            ws_summary.column_dimensions[col_letter].width = width
        
        # 保存文件
        output_file = "PMC排产Sep01-Sep07订单ROI分析_含供应商汇总.xlsx"
        book.save(output_file)
        print(f"\n已成功添加供应商缺料汇总表，保存为: {output_file}")
        
        # 显示前10名供应商
        print(f"\n缺料金额前10名供应商:")
        top10 = supplier_summary.head(10)
        for idx, row in top10.iterrows():
            print(f"{idx+1:2d}. {row['供应商名称']:15s} ¥{row['缺料金额合计(RMB)']:>12,.2f} ({row['占总缺料比例(%)']:>6.2f}%)")
        
        return output_file
        
    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    result = add_supplier_shortage_summary()
    if result:
        print(f"\n任务完成! 新文件: {result}")